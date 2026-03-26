"""
Chaney Place Townhomes — BI Dashboard
Reads the redIQ underwriting model and displays key metrics, charts, and tables.
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from pathlib import Path
import datetime

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = Path(__file__).parent / "Chaney_Place_Townhomes_0757_UNPROTECTED.xlsm"

st.set_page_config(
    page_title="Chaney Place Townhomes — BI Dashboard",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

@st.cache_data
def load_workbook():
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)


def safe_val(ws, coord, default=None):
    v = ws[coord].value
    return v if v is not None else default


def safe_float(ws, coord, default=0.0):
    v = ws[coord].value
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def row_floats(ws, row, start_col, count):
    """Read `count` float values from a row starting at start_col."""
    return [safe_float(ws, ws.cell(row=row, column=start_col + i).coordinate, 0) for i in range(count)]


# ---------------------------------------------------------------------------
# Data loading functions
# ---------------------------------------------------------------------------

@st.cache_data
def load_deal_overview():
    wb = load_workbook()
    ws_input = wb["Input"]
    ws_summary = wb["Summary"]

    # Input sheet: total units at I12, but that cell had value 1 (bug).
    # Derive from unit mix: 63 x 2BR + 17 x 3BR = 80
    total_units = 80
    purchase_price = safe_float(ws_summary, "M7", 13009400)

    deal = {
        "name": safe_val(ws_input, "E10", "Chaney Place Townhomes"),
        "address": safe_val(ws_summary, "E7", "1060 Southeast Chaney Place Drive"),
        "city_state": safe_val(ws_input, "M4", "Huntsville, AL"),
        "total_units": total_units,
        "closing_date": safe_val(ws_input, "M10", ""),
        "hold_period": safe_val(ws_input, "M11", 7),
        "year_built": 2014,
        "purchase_price": purchase_price,
        "price_per_unit": purchase_price / total_units,
        "unlev_irr": safe_float(ws_summary, "O7", 0),
        "unlev_em": safe_float(ws_summary, "P7", 0),
        "lev_irr": safe_float(ws_summary, "O8", 0),
        "lev_em": safe_float(ws_summary, "P8", 0),
        "going_in_cap": 0.055,
        "exit_cap": 0.055,
    }

    return deal


@st.cache_data
def load_property_cf():
    """Property CF sheet — annual projected cash flows."""
    wb = load_workbook()
    ws = wb["Property CF"]

    # Years in row 4, cols C(3) to N(14): T12 Adj + Years 1-11
    years = []
    year_labels = []
    for c in range(3, 15):
        v = ws.cell(row=4, column=c).value
        if v is not None:
            if isinstance(v, datetime.datetime):
                years.append(c)
                year_labels.append(v.strftime("%b %Y"))
            else:
                years.append(c)
                year_labels.append(str(v))

    n = len(years)

    # Correct row mapping (verified from scan)
    row_map = {
        "Potential Market Rent": 10,
        "Loss to Lease": 12,
        "Gross Potential Revenue": 13,
        "Vacancy": 15,
        "Concessions": 17,
        "Collection Loss / Bad Debt": 19,
        "Base Rental Revenue": 20,
        "Other Income": 25,
        "Effective Gross Revenue": 27,
        "Total Operating Expenses": 47,
        "NOI (bef. Reserves)": 52,
        "NOI (aft. Reserves)": 55,
        "Operating Cash Flow": 69,
    }

    data = {}
    for label, row in row_map.items():
        data[label] = [safe_float(ws, ws.cell(row=row, column=c).coordinate, 0) for c in years]

    # Operating metrics
    metrics_map = {
        "Physical Occupancy": 90,
        "Economic Occupancy": 91,
        "OpEx Margin": 92,
        "NOI Yield": 94,
    }
    metrics = {}
    for label, row in metrics_map.items():
        metrics[label] = [safe_float(ws, ws.cell(row=row, column=c).coordinate, 0) for c in years]

    return year_labels, data, metrics


@st.cache_data
def load_investment_cf():
    """Cash Flow sheet — investment-level cash flows."""
    wb = load_workbook()
    ws = wb["Cash Flow"]

    # Year labels in row 5, cols D(4) to N(14)
    cols = []
    year_labels = []
    for c in range(4, 15):
        v = ws.cell(row=5, column=c).value
        if v is not None:
            cols.append(c)
            year_labels.append(int(v) if isinstance(v, (int, float)) else str(v))

    n = len(cols)

    # Correct row mapping
    row_map = {
        "Acquisition Cost": 13,
        "Effective Gross Revenue": 18,
        "Operating Expenses": 19,
        "NOI": 21,
        "Operating Cash Flow": 25,
        "Net Sales Proceeds": 32,
        "Unleveraged Cash Flow": 36,
        "Debt Service": 42,
        "Leveraged Cash Flow": 57,
    }

    data = {}
    for label, row in row_map.items():
        data[label] = [safe_float(ws, ws.cell(row=row, column=c).coordinate, 0) for c in cols]

    # Cash on Cash
    data["Cash on Cash"] = [safe_float(ws, ws.cell(row=62, column=c).coordinate, 0) for c in cols]

    return year_labels, data


@st.cache_data
def load_historical_cf():
    """Historical CF sheet — trailing periods."""
    wb = load_workbook()
    ws = wb["Historical CF"]

    periods = ["T12", "T9", "T6", "T3", "T1", "Pro Forma Yr1"]
    period_cols = [4, 7, 10, 13, 16, 20]

    # Correct row mapping
    row_map = {
        "Potential Market Rent": 9,
        "Loss to Lease": 10,
        "Gross Potential Revenue": 11,
        "Vacancy": 13,
        "Concessions": 14,
        "Non-Revenue Units": 15,
        "Collection Loss / Bad Debt": 16,
        "Base Rental Revenue": 17,
        "Expense Reimbursements": 21,
        "Other Residential Income": 22,
        "Other Income": 24,
        "Effective Gross Revenue": 26,
        "Repair & Maintenance": 31,
        "Landscaping / Grounds": 34,
        "Personnel": 35,
        "Marketing / Advertising": 36,
        "Administrative": 38,
        "Electricity": 40,
        "Other Utilities": 43,
        "Insurance": 44,
        "Real Estate Taxes": 45,
        "Property Management Fee": 47,
        "Total Operating Expenses": 51,
        "NOI (bef. Reserves)": 56,
        "NOI (aft. Reserves)": 59,
    }

    data = {}
    for label, row in row_map.items():
        data[label] = [safe_float(ws, ws.cell(row=row, column=col).coordinate, 0) for col in period_cols]

    return periods, data


@st.cache_data
def load_t12_monthly():
    """T12 sheet — monthly GL detail."""
    wb = load_workbook()
    ws = wb["T12"]

    months = []
    for c in range(4, 16):
        v = ws.cell(row=3, column=c).value
        if v is not None:
            if isinstance(v, datetime.datetime):
                months.append(v.strftime("%b %Y"))
            else:
                months.append(str(v))

    n = len(months)

    # Correct row mapping
    row_map = {
        "Gross Potential Rent": 9,
        "Vacancy": 11,
        "Concessions": 12,
        "Bad Debt": 14,
        "Total Vacancy/Concessions/Loss": 15,
        "Net Rental Income": 16,
        "Total Other Income": 35,
        "Total Income": 36,
        "Personnel": 45,
        "Management Fees": 48,
        "Total Contract Services": 56,
        "Administrative": 77,
        "Marketing": 85,
        "Utilities": 89,
        "Maintenance & Repairs": 110,
        "Taxes & Insurance": 114,
        "Total Operating Expenses": 115,
        "Net Operating Income": 116,
    }

    data = {}
    for label, row in row_map.items():
        data[label] = [safe_float(ws, ws.cell(row=row, column=4 + i).coordinate, 0) for i in range(n)]

    return months, data


@st.cache_data
def load_sources_uses():
    """Sources & Uses sheet."""
    wb = load_workbook()
    ws = wb["Sources & Uses"]

    sources = []
    for r in range(9, 20):
        label = ws.cell(row=r, column=2).value
        amt = ws.cell(row=r, column=3).value
        if label and amt and "TOTAL" not in str(label).upper():
            try:
                sources.append({"Item": str(label), "Amount": float(amt)})
            except (TypeError, ValueError):
                pass

    uses = []
    seen = set()
    for r in range(10, 45):
        label = ws.cell(row=r, column=7).value
        amt = ws.cell(row=r, column=8).value
        if label and amt:
            key = str(label).strip()
            if "TOTAL" in key.upper() or key in seen:
                continue
            seen.add(key)
            try:
                uses.append({"Item": key, "Amount": float(amt)})
            except (TypeError, ValueError):
                pass

    return sources, uses


@st.cache_data
def load_sensitivity():
    """Sensitivity Analysis sheet."""
    wb = load_workbook()
    ws = wb["Sensitivity Analysis"]

    # Purchase price sensitivity — row 7 headers, cols D-H
    pp_headers = []
    for c in range(4, 9):
        v = ws.cell(row=7, column=c).value
        if v is not None:
            try:
                pp_headers.append(f"${float(v):,.0f}")
            except (TypeError, ValueError):
                pp_headers.append(str(v))

    pp_rows = {}
    pp_labels = {
        8: "Per Unit",
        9: "Per SF",
        10: "Going-In Cap Rate",
        12: "Unleveraged IRR",
        13: "Leveraged IRR",
        14: "Equity Multiple",
    }
    for row_num, label in pp_labels.items():
        pp_rows[label] = [safe_float(ws, ws.cell(row=row_num, column=c).coordinate, 0) for c in range(4, 9)]

    # Cap rate sensitivity — cols L-P
    cap_headers = []
    for c in range(12, 17):
        v = ws.cell(row=7, column=c).value
        if v is not None:
            try:
                cap_headers.append(f"{float(v):.2%}")
            except (TypeError, ValueError):
                cap_headers.append(str(v))

    cap_rows = {}
    for row_num, label in pp_labels.items():
        cap_rows[label] = [safe_float(ws, ws.cell(row=row_num, column=c).coordinate, 0) for c in range(12, 17)]

    return pp_headers, pp_rows, cap_headers, cap_rows


# ---------------------------------------------------------------------------
# LAYOUT
# ---------------------------------------------------------------------------

with st.sidebar:
    st.title("Navigation")
    page = st.radio(
        "Select View",
        [
            "Deal Overview",
            "Property Cash Flow",
            "Historical Performance",
            "Monthly T12 Detail",
            "Sources & Uses",
            "Sensitivity Analysis",
        ],
    )

# Header
deal = load_deal_overview()
st.title(deal["name"])
st.caption(f"{deal['address']} | {deal['city_state']}")

# ==========================================================================
# PAGE: Deal Overview
# ==========================================================================
if page == "Deal Overview":
    st.header("Investment Summary")

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Purchase Price", f"${deal['purchase_price']:,.0f}")
    col2.metric("Price / Unit", f"${deal['price_per_unit']:,.0f}")
    col3.metric("Total Units", f"{deal['total_units']}")
    col4.metric("Hold Period", f"{deal['hold_period']} years")
    col5.metric("Year Built", f"{deal['year_built']}")

    st.divider()

    col1, col2, col3, col4, col5, col6 = st.columns(6)
    col1.metric("Unleveraged IRR", f"{deal['unlev_irr']:.2%}")
    col2.metric("Equity Multiple", f"{deal['unlev_em']:.2f}x")
    col3.metric("Leveraged IRR", f"{deal['lev_irr']:.2%}")
    col4.metric("Leveraged EM", f"{deal['lev_em']:.2f}x")
    col5.metric("Going-In Cap", f"{deal['going_in_cap']:.2%}")
    col6.metric("Exit Cap", f"{deal['exit_cap']:.2%}")

    st.divider()

    # NOI chart
    year_labels, pcf, metrics = load_property_cf()
    noi_vals = pcf.get("NOI (bef. Reserves)", [])
    if year_labels and noi_vals:
        st.subheader("Projected NOI Over Hold Period")
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=year_labels,
            y=noi_vals,
            marker_color="#2563EB",
            text=[f"${v:,.0f}" for v in noi_vals],
            textposition="outside",
        ))
        fig.update_layout(
            yaxis_title="NOI ($)",
            yaxis_tickformat="$,.0f",
            height=400,
            margin=dict(t=20),
        )
        st.plotly_chart(fig, use_container_width=True)

    # Investment Cash Flow waterfall
    inv_labels, inv_cf = load_investment_cf()
    unlev = inv_cf.get("Unleveraged Cash Flow", [])
    if inv_labels and unlev:
        st.subheader("Unleveraged Cash Flow by Year")
        colors = ["#EF4444" if v < 0 else "#10B981" for v in unlev]
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=[f"Year {y}" for y in inv_labels],
            y=unlev,
            marker_color=colors,
            text=[f"${v:,.0f}" for v in unlev],
            textposition="outside",
        ))
        fig.update_layout(
            yaxis_title="Cash Flow ($)",
            yaxis_tickformat="$,.0f",
            height=400,
            margin=dict(t=20),
        )
        st.plotly_chart(fig, use_container_width=True)

# ==========================================================================
# PAGE: Property Cash Flow
# ==========================================================================
elif page == "Property Cash Flow":
    st.header("Property Cash Flow Projections")

    year_labels, pcf, metrics = load_property_cf()
    if not year_labels:
        st.warning("No property cash flow data found.")
    else:
        # Revenue vs Expense vs NOI
        st.subheader("Revenue, Expenses & NOI")
        egr = pcf.get("Effective Gross Revenue", [0] * len(year_labels))
        opex = [abs(v) for v in pcf.get("Total Operating Expenses", [0] * len(year_labels))]
        noi = pcf.get("NOI (bef. Reserves)", [0] * len(year_labels))

        fig = go.Figure()
        fig.add_trace(go.Bar(name="Eff. Gross Revenue", x=year_labels, y=egr, marker_color="#3B82F6"))
        fig.add_trace(go.Bar(name="Operating Expenses", x=year_labels, y=opex, marker_color="#F59E0B"))
        fig.add_trace(go.Scatter(
            name="NOI", x=year_labels, y=noi,
            mode="lines+markers", line=dict(color="#10B981", width=3), marker=dict(size=8),
        ))
        fig.update_layout(
            barmode="group", yaxis_tickformat="$,.0f", height=450,
            margin=dict(t=20), legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        st.plotly_chart(fig, use_container_width=True)

        # NOI Margin & Occupancy
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("NOI Margin (%)")
            margins = []
            for e, n in zip(egr, noi):
                try:
                    margins.append(n / e * 100 if e else 0)
                except (TypeError, ZeroDivisionError):
                    margins.append(0)
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(
                x=year_labels, y=margins,
                mode="lines+markers+text",
                text=[f"{m:.1f}%" for m in margins],
                textposition="top center",
                line=dict(color="#8B5CF6", width=3), marker=dict(size=8),
            ))
            fig2.update_layout(yaxis_title="NOI Margin (%)", height=350, margin=dict(t=20))
            st.plotly_chart(fig2, use_container_width=True)

        with col2:
            st.subheader("Occupancy")
            phys_occ = [v * 100 for v in metrics.get("Physical Occupancy", [0] * len(year_labels))]
            econ_occ = [v * 100 for v in metrics.get("Economic Occupancy", [0] * len(year_labels))]
            fig3 = go.Figure()
            fig3.add_trace(go.Scatter(
                name="Physical Occupancy", x=year_labels, y=phys_occ,
                mode="lines+markers", line=dict(color="#3B82F6", width=2),
            ))
            fig3.add_trace(go.Scatter(
                name="Economic Occupancy", x=year_labels, y=econ_occ,
                mode="lines+markers", line=dict(color="#10B981", width=2),
            ))
            fig3.update_layout(
                yaxis_title="Occupancy (%)", height=350, margin=dict(t=20),
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
            )
            st.plotly_chart(fig3, use_container_width=True)

        # Data table
        st.subheader("Detailed Cash Flow Table")
        df = pd.DataFrame(pcf, index=year_labels).T
        st.dataframe(df.style.format("${:,.0f}"), use_container_width=True, height=500)

# ==========================================================================
# PAGE: Historical Performance
# ==========================================================================
elif page == "Historical Performance":
    st.header("Historical vs. Pro Forma Performance")

    periods, hist = load_historical_cf()
    if not periods:
        st.warning("No historical data found.")
    else:
        # Key metrics comparison
        st.subheader("EGR, OpEx & NOI by Trailing Period")
        egr = hist.get("Effective Gross Revenue", [0] * len(periods))
        opex = hist.get("Total Operating Expenses", [0] * len(periods))
        noi = hist.get("NOI (bef. Reserves)", [0] * len(periods))

        fig = go.Figure()
        fig.add_trace(go.Bar(name="EGR", x=periods, y=egr, marker_color="#3B82F6"))
        fig.add_trace(go.Bar(name="OpEx", x=periods, y=opex, marker_color="#F59E0B"))
        fig.add_trace(go.Bar(name="NOI", x=periods, y=noi, marker_color="#10B981"))
        fig.update_layout(
            barmode="group", yaxis_tickformat="$,.0f", height=450,
            margin=dict(t=20), legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        st.plotly_chart(fig, use_container_width=True)

        # Revenue waterfall
        st.subheader("Revenue Breakdown by Period")
        rev_items = {
            "Potential Market Rent": "#3B82F6",
            "Loss to Lease": "#EF4444",
            "Other Income": "#8B5CF6",
        }
        fig2 = go.Figure()
        for item, color in rev_items.items():
            if item in hist:
                fig2.add_trace(go.Bar(name=item, x=periods, y=hist[item], marker_color=color))
        fig2.update_layout(
            barmode="group", yaxis_tickformat="$,.0f", height=400,
            margin=dict(t=20), legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        st.plotly_chart(fig2, use_container_width=True)

        # Expense detail: T12 vs Pro Forma
        st.subheader("Expense Detail: T12 Actual vs Pro Forma Year 1")
        expense_items = [
            "Repair & Maintenance", "Landscaping / Grounds", "Personnel",
            "Marketing / Advertising", "Administrative", "Electricity",
            "Other Utilities", "Insurance", "Real Estate Taxes",
            "Property Management Fee",
        ]
        t12_vals = []
        pf1_vals = []
        labels = []
        for item in expense_items:
            if item in hist:
                t12_v = hist[item][0]  # T12
                pf1_v = hist[item][5]  # Pro Forma Yr1
                if t12_v != 0 or pf1_v != 0:
                    labels.append(item)
                    t12_vals.append(t12_v)
                    pf1_vals.append(pf1_v)

        if labels:
            fig3 = go.Figure()
            fig3.add_trace(go.Bar(name="T12 Actual", x=labels, y=t12_vals, marker_color="#F59E0B"))
            fig3.add_trace(go.Bar(name="Pro Forma Yr1", x=labels, y=pf1_vals, marker_color="#3B82F6"))
            fig3.update_layout(
                barmode="group", yaxis_tickformat="$,.0f", height=450,
                margin=dict(t=20), xaxis_tickangle=-45,
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
            )
            st.plotly_chart(fig3, use_container_width=True)

        # Full table
        st.subheader("Full Historical Data")
        df = pd.DataFrame(hist, index=periods).T
        st.dataframe(df.style.format("${:,.0f}"), use_container_width=True, height=600)

# ==========================================================================
# PAGE: Monthly T12 Detail
# ==========================================================================
elif page == "Monthly T12 Detail":
    st.header("Monthly T12 Actuals (Dec 2024 - Nov 2025)")

    months, t12 = load_t12_monthly()
    if not months:
        st.warning("No T12 monthly data found.")
    else:
        # NOI trend
        st.subheader("Monthly NOI Trend")
        noi = t12.get("Net Operating Income", [0] * len(months))
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=months, y=noi, mode="lines+markers",
            line=dict(color="#10B981", width=3), marker=dict(size=8),
            fill="tozeroy", fillcolor="rgba(16, 185, 129, 0.1)",
        ))
        fig.update_layout(yaxis_title="NOI ($)", yaxis_tickformat="$,.0f", height=400, margin=dict(t=20))
        st.plotly_chart(fig, use_container_width=True)

        # Revenue vs Expenses monthly
        st.subheader("Monthly Total Income vs Operating Expenses")
        total_inc = t12.get("Total Income", [0] * len(months))
        opex = t12.get("Total Operating Expenses", [0] * len(months))

        fig2 = go.Figure()
        fig2.add_trace(go.Bar(name="Total Income", x=months, y=total_inc, marker_color="#3B82F6"))
        fig2.add_trace(go.Bar(name="OpEx", x=months, y=opex, marker_color="#F59E0B"))
        fig2.update_layout(
            barmode="group", yaxis_tickformat="$,.0f", height=400,
            margin=dict(t=20), legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        st.plotly_chart(fig2, use_container_width=True)

        # Vacancy impact
        st.subheader("Gross Potential Rent vs Net Rental Income")
        gpr = t12.get("Gross Potential Rent", [0] * len(months))
        nri = t12.get("Net Rental Income", [0] * len(months))

        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(
            name="Gross Potential Rent", x=months, y=gpr,
            mode="lines+markers", line=dict(color="#3B82F6", width=2),
        ))
        fig3.add_trace(go.Scatter(
            name="Net Rental Income", x=months, y=nri,
            mode="lines+markers", line=dict(color="#10B981", width=2),
        ))
        fig3.update_layout(
            yaxis_tickformat="$,.0f", height=400, margin=dict(t=20),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        st.plotly_chart(fig3, use_container_width=True)

        # Expense breakdown
        st.subheader("Monthly Expense Breakdown")
        exp_items = ["Personnel", "Management Fees", "Total Contract Services",
                     "Administrative", "Marketing", "Utilities",
                     "Maintenance & Repairs", "Taxes & Insurance"]
        exp_colors = ["#EF4444", "#F59E0B", "#3B82F6", "#8B5CF6",
                      "#EC4899", "#10B981", "#F97316", "#6366F1"]
        fig4 = go.Figure()
        for item, color in zip(exp_items, exp_colors):
            if item in t12:
                fig4.add_trace(go.Bar(name=item, x=months, y=t12[item], marker_color=color))
        fig4.update_layout(
            barmode="stack", yaxis_tickformat="$,.0f", height=500,
            margin=dict(t=20), legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        st.plotly_chart(fig4, use_container_width=True)

        # Full data table
        st.subheader("Monthly Data Table")
        df = pd.DataFrame(t12, index=months).T
        st.dataframe(df.style.format("${:,.0f}"), use_container_width=True, height=500)

# ==========================================================================
# PAGE: Sources & Uses
# ==========================================================================
elif page == "Sources & Uses":
    st.header("Sources & Uses at Closing")

    sources, uses = load_sources_uses()

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Sources")
        if sources:
            df_s = pd.DataFrame(sources)
            st.dataframe(df_s.style.format({"Amount": "${:,.0f}"}), use_container_width=True, hide_index=True)
            fig = go.Figure(go.Pie(
                labels=[s["Item"] for s in sources],
                values=[s["Amount"] for s in sources],
                marker_colors=["#3B82F6", "#10B981", "#F59E0B", "#8B5CF6"],
                hole=0.4,
            ))
            fig.update_layout(height=350, margin=dict(t=20))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No sources data found.")

    with col2:
        st.subheader("Uses")
        if uses:
            df_u = pd.DataFrame(uses)
            st.dataframe(df_u.style.format({"Amount": "${:,.0f}"}), use_container_width=True, hide_index=True)
            fig = go.Figure(go.Pie(
                labels=[u["Item"] for u in uses],
                values=[u["Amount"] for u in uses],
                marker_colors=["#EF4444", "#F59E0B", "#3B82F6", "#8B5CF6", "#10B981"],
                hole=0.4,
            ))
            fig.update_layout(height=350, margin=dict(t=20))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No uses data found.")

    total_sources = sum(s["Amount"] for s in sources) if sources else 0
    total_uses = sum(u["Amount"] for u in uses) if uses else 0
    st.divider()
    col1, col2, col3 = st.columns(3)
    col1.metric("Total Sources", f"${total_sources:,.0f}")
    col2.metric("Total Uses", f"${total_uses:,.0f}")
    col3.metric("Difference", f"${total_sources - total_uses:,.0f}")

# ==========================================================================
# PAGE: Sensitivity Analysis
# ==========================================================================
elif page == "Sensitivity Analysis":
    st.header("Sensitivity Analysis")

    pp_headers, pp_rows, cap_headers, cap_rows = load_sensitivity()

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Purchase Price Sensitivity")
        if pp_headers and pp_rows:
            df_pp = pd.DataFrame(pp_rows, index=pp_headers).T
            st.dataframe(df_pp, use_container_width=True)

            if "Unleveraged IRR" in pp_rows:
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=pp_headers,
                    y=[v * 100 for v in pp_rows["Unleveraged IRR"]],
                    mode="lines+markers+text",
                    text=[f"{v:.2%}" for v in pp_rows["Unleveraged IRR"]],
                    textposition="top center",
                    line=dict(color="#3B82F6", width=3), marker=dict(size=10),
                ))
                fig.update_layout(
                    title="Unleveraged IRR vs Purchase Price",
                    yaxis_title="IRR (%)", height=400, margin=dict(t=40),
                )
                st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Cap Rate Sensitivity")
        if cap_headers and cap_rows:
            df_cap = pd.DataFrame(cap_rows, index=cap_headers).T
            st.dataframe(df_cap, use_container_width=True)

            if "Unleveraged IRR" in cap_rows:
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=cap_headers,
                    y=[v * 100 for v in cap_rows["Unleveraged IRR"]],
                    mode="lines+markers+text",
                    text=[f"{v:.2%}" for v in cap_rows["Unleveraged IRR"]],
                    textposition="top center",
                    line=dict(color="#10B981", width=3), marker=dict(size=10),
                ))
                fig.update_layout(
                    title="Unleveraged IRR vs Exit Cap Rate",
                    yaxis_title="IRR (%)", height=400, margin=dict(t=40),
                )
                st.plotly_chart(fig, use_container_width=True)

# Footer
st.divider()
st.caption("Data sourced from redIQ underwriting model | Chaney Place Townhomes, Huntsville, AL")
